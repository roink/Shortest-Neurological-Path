import numpy as np
from brain_path import BrainPath
from paths import Paths
from heapq import heappush
from heapq import heappop

distances = np.loadtxt('/home/philipp/alzheimers/combined-atlas/delaunay_triangulation/delaunay-edges.txt',
                       delimiter=',')

[m, _] = np.shape(distances)

from openpyxl import load_workbook

wb = load_workbook(
    '/home/philipp/alzheimers/neurodegeneration-forecast/Neurodeg_Data_SUVR/SUVRs_New_Atlas_Modeling.xlsx')
ws = wb['SUVRs_New_Atlas']
concentrations = np.array([[i.value for i in j] for j in ws['B2':'UZ21']])
concentrations = np.mean(concentrations, axis=0)

heap = []

k = 10

with open('10-least-tau-path.txt', 'w') as file:
    for start in range(m):
        print(start)

        ps = Paths(k, m)

        p = BrainPath(start, concentrations[start])
        heappush(heap, p)

        while len(heap) > 0:
            path: BrainPath = heappop(heap)
            for next_node in np.nonzero(distances[path.last()])[0]:
                if next_node not in path.elements:
                    new_path = path.add(next_node, concentrations[next_node])
                    if ps.add(new_path):
                        heappush(heap, new_path)
        file.write(ps.str(start) + '\n')
